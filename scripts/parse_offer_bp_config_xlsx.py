#!/usr/bin/env python3
"""
Parse Offer BP definition Excel exports for PM-facing structure discovery.

Lists worksheets and the first rows of each sheet (step names, approvals, branches).
Requires openpyxl: pip install openpyxl

Usage:
  python3 scripts/parse_offer_bp_config_xlsx.py research/Offer/bp-config-samples/
"""

from __future__ import annotations

import glob
import os
import sys


def main() -> int:
    root = sys.argv[1] if len(sys.argv) > 1 else os.path.join("research", "Offer", "bp-config-samples")
    paths = sorted(glob.glob(os.path.join(root, "*.xlsx")))
    if not paths:
        print(f"No .xlsx files under {root}.")
        print("Copy exports here, e.g. offer_default_definition.xlsx and vmware_offer_default_definition.xlsx.")
        return 0
    try:
        import openpyxl  # type: ignore
    except ImportError:
        print("Missing dependency: pip install openpyxl")
        return 1

    for path in paths:
        print("=" * 72)
        print("FILE", path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        print("  sheets:", ", ".join(wb.sheetnames))
        for name in wb.sheetnames:
            ws = wb[name]
            print(f"  --- {name} (preview) ---")
            for i, row in enumerate(ws.iter_rows(max_row=25, max_col=10, values_only=True)):
                if i >= 12:
                    break
                print("   ", row)
        wb.close()

    print()
    print("Hypothesis prompts for Pharos (validate before charting):")
    print("  • Extra approval steps in config → step volume / duration for exact task_name in bp_event_record_stats")
    print("  • Writer-generated review path → share of offers hitting Review Writer Generated Document")
    print("  • Parallel approvers → distinct assignee patterns (needs stable join; not assumed available)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
